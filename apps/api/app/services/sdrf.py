from __future__ import annotations

import csv
import io
import json
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
except Exception:  # pragma: no cover - optional in minimal installs
    Workbook = None
    load_workbook = None

try:
    import xlrd
except Exception:  # pragma: no cover - optional in minimal installs
    xlrd = None


REQUIRED_COLUMNS = [
    "source name",
    "characteristics[organism]",
    "characteristics[organism part]",
    "characteristics[disease]",
    "characteristics[biological replicate]",
    "assay name",
    "technology type",
    "comment[proteomics data acquisition method]",
    "comment[label]",
    "comment[instrument]",
    "comment[cleavage agent details]",
    "comment[fraction identifier]",
    "comment[technical replicate]",
    "comment[data file]",
    "factor value[disease]",
]

TEMPLATE_REGISTRY: dict[str, dict[str, Any]] = {
    "human": {
        "version": "v1.1.0",
        "columns": [
            "characteristics[organism]",
            "characteristics[disease]",
            "characteristics[organism part]",
            "characteristics[age]",
            "characteristics[sex]",
        ],
    },
    "vertebrates": {
        "version": "v1.1.0",
        "columns": [
            "characteristics[organism]",
            "characteristics[disease]",
            "characteristics[developmental stage]",
            "characteristics[strain]",
        ],
    },
    "dia": {
        "version": "v1.0.0",
        "columns": [
            "comment[proteomics data acquisition method]",
            "comment[isolation window]",
            "comment[scan window lower limit]",
            "comment[scan window upper limit]",
        ],
    },
    "labeling": {
        "version": "v1.0.0",
        "columns": ["comment[label]", "comment[modification parameters]"],
    },
    "affinity-proteomics": {
        "version": "v1.0.0",
        "columns": ["comment[assay kit]", "comment[technology platform]"],
    },
    "single-cell": {
        "version": "v1.0.0",
        "columns": ["characteristics[cell type]", "comment[carrier proteome]"],
    },
}


def detect_delimiter(text: str) -> str:
    sample = text[:4096]
    try:
        return csv.Sniffer().sniff(sample, delimiters="\t,;").delimiter
    except csv.Error:
        return "\t" if "\t" in sample else ","


def normalize_header(header: str) -> str:
    return header.strip().lstrip("\ufeff").removeprefix("ï»¿").strip()


def canonical_sdrf_header(header: str) -> str:
    normalized = normalize_header(header)
    lower = normalized.lower()
    if lower == "source name":
        return "source name"
    if lower == "assay name":
        return "assay name"
    if lower == "material type":
        return "material type"
    for prefix in ("characteristics[", "comment[", "factor value["):
        if lower.startswith(prefix):
            return lower
    return normalized


def normalize_sdrf_table(headers: list[str], rows: list[dict[str, Any]]) -> tuple[list[str], list[dict[str, Any]]]:
    canonical_headers: list[str] = []
    canonical_for_header = {header: canonical_sdrf_header(header) for header in headers}
    for header in headers:
        canonical = canonical_for_header[header]
        if canonical not in canonical_headers:
            canonical_headers.append(canonical)

    canonical_rows: list[dict[str, Any]] = []
    for row in rows:
        next_row: dict[str, Any] = {}
        for header in headers:
            canonical = canonical_for_header[header]
            value = row.get(header, "")
            existing = next_row.get(canonical, "")
            if str(existing).strip():
                continue
            next_row[canonical] = value
        for key, value in row.items():
            if key in canonical_for_header:
                continue
            canonical = canonical_sdrf_header(key)
            if canonical not in canonical_headers:
                canonical_headers.append(canonical)
            existing = next_row.get(canonical, "")
            if not str(existing).strip():
                next_row[canonical] = value
        canonical_rows.append(next_row)
    return canonical_headers, canonical_rows


def rows_to_table(headers: list[str], matrix: list[list[str]]) -> dict[str, Any]:
    headers = [normalize_header(header) for header in headers]
    rows = []
    for values in matrix:
        padded = values + [""] * max(0, len(headers) - len(values))
        rows.append({header: padded[index] if index < len(padded) else "" for index, header in enumerate(headers)})
    headers, rows = normalize_sdrf_table(headers, rows)
    return {
        "headers": headers,
        "rows": rows,
        "column_metadata": classify_columns(headers),
        "dirty": False,
        "validation_state": {},
    }


def parse_text_table(content: str) -> dict[str, Any]:
    delimiter = detect_delimiter(content)
    reader = csv.reader(io.StringIO(content), delimiter=delimiter)
    rows = [row for row in reader if any(cell.strip() for cell in row)]
    if not rows:
        return {"headers": [], "rows": [], "column_metadata": {}, "dirty": False, "validation_state": {}}
    return rows_to_table([cell.strip() for cell in rows[0]], rows[1:])


def parse_file(path: Path) -> dict[str, Any]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xlx"}:
        if load_workbook is None:
            raise RuntimeError("openpyxl is required to parse Excel files")
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        values = [[cell if cell is not None else "" for cell in row] for row in sheet.iter_rows(values_only=True)]
        if not values:
            return {"headers": [], "rows": [], "column_metadata": {}, "dirty": False, "validation_state": {}}
        return rows_to_table([str(cell).strip() for cell in values[0]], [[str(cell) for cell in row] for row in values[1:]])
    if suffix == ".xls":
        if xlrd is None:
            raise RuntimeError("xlrd is required to parse .xls files")
        workbook = xlrd.open_workbook(str(path))
        sheet = workbook.sheet_by_index(0)
        values = [[sheet.cell_value(row, col) for col in range(sheet.ncols)] for row in range(sheet.nrows)]
        if not values:
            return {"headers": [], "rows": [], "column_metadata": {}, "dirty": False, "validation_state": {}}
        return rows_to_table([str(cell).strip() for cell in values[0]], [[str(cell) for cell in row] for row in values[1:]])
    return parse_text_table(path.read_text(encoding="utf-8", errors="replace"))


def classify_columns(headers: list[str]) -> dict[str, Any]:
    metadata: dict[str, Any] = {}
    for header in headers:
        lower = header.lower()
        if lower == "source name":
            section = "sample"
        elif lower.startswith("characteristics["):
            section = "sample"
        elif lower.startswith("comment["):
            section = "data_file"
        elif lower.startswith("factor value["):
            section = "factor"
        elif lower == "assay name":
            section = "assay"
        else:
            section = "other"
        metadata[header] = {"section": section, "required": header in REQUIRED_COLUMNS}
    return metadata


def table_to_tsv(headers: list[str], rows: list[dict[str, Any]]) -> str:
    output = io.StringIO()
    writer = csv.writer(output, delimiter="\t", lineterminator="\n")
    writer.writerow(headers)
    for row in rows:
        writer.writerow([row.get(header, "") for header in headers])
    return output.getvalue()


def write_xlsx(path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is required to export Excel files")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SDRF"
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    workbook.save(path)


def validate_structural(headers: list[str], rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    header_set = set(headers)
    for column in REQUIRED_COLUMNS:
        if column not in header_set:
            issues.append({
                "severity": "error",
                "message": f"Required SDRF column is missing: {column}",
                "column": column,
                "rule": "required-column",
                "suggested_fix": f"Add {column} before export.",
            })

    for index, row in enumerate(rows):
        for column in ["source name", "assay name", "comment[data file]"]:
            if column in header_set and not str(row.get(column, "")).strip():
                issues.append({
                    "severity": "error",
                    "message": f"{column} is empty.",
                    "row": index,
                    "column": column,
                    "rule": "required-value",
                    "suggested_fix": f"Fill {column} for this row.",
                })
    for header in headers:
        if header != header.strip():
            issues.append({
                "severity": "warning",
                "message": f"Column header has leading or trailing whitespace: {header!r}",
                "column": header,
                "rule": "header-whitespace",
                "suggested_fix": "Trim the column header.",
            })
        lower = header.lower()
        if any(lower.startswith(prefix) for prefix in ["characteristics[", "comment[", "factor value["]) and not header.endswith("]"):
            issues.append({
                "severity": "error",
                "message": f"Column header is missing a closing bracket: {header}",
                "column": header,
                "rule": "header-format",
                "suggested_fix": "Use the SDRF header format prefix[term].",
            })
    if not rows:
        issues.append({
            "severity": "warning",
            "message": "The SDRF table has no data rows.",
            "rule": "empty-table",
            "suggested_fix": "Add at least one sample-file row.",
        })
    return issues


def find_sdrf_pipeline_executable() -> str | None:
    for command in ("sdrf", "sdrf-pipelines"):
        executable = shutil.which(command)
        if executable:
            return executable
    return None


def validate_with_sdrf_pipelines(headers: list[str], rows: list[dict[str, Any]]) -> list[dict[str, Any]] | None:
    executable = find_sdrf_pipeline_executable()
    if not executable:
        return None
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "candidate.sdrf.tsv"
        path.write_text(table_to_tsv(headers, rows), encoding="utf-8")
        result = subprocess.run(
            [executable, "validate", "-s", str(path)],
            capture_output=True,
            text=True,
            timeout=60,
            check=False,
        )
        if result.returncode == 0:
            return []
        lines = [line.strip() for line in (result.stdout + "\n" + result.stderr).splitlines() if line.strip()]
        return [
            {
                "severity": "error",
                "message": line,
                "rule": "sdrf-pipelines",
                "suggested_fix": "Review this validator message and update the SDRF table.",
            }
            for line in lines
        ]


def validate_table(headers: list[str], rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    headers, rows = normalize_sdrf_table(headers, rows)
    pipeline_issues = validate_with_sdrf_pipelines(headers, rows)
    issues = pipeline_issues if pipeline_issues is not None else validate_structural(headers, rows)
    summary = {
        "errors": sum(1 for issue in issues if issue["severity"] == "error"),
        "warnings": sum(1 for issue in issues if issue["severity"] == "warning"),
        "infos": sum(1 for issue in issues if issue["severity"] == "info"),
        "validator": "sdrf-pipelines" if pipeline_issues is not None else "structural-fallback",
    }
    return issues, summary


def default_table() -> dict[str, Any]:
    return {
        "headers": REQUIRED_COLUMNS[:],
        "rows": [],
        "column_metadata": classify_columns(REQUIRED_COLUMNS),
        "dirty": False,
        "validation_state": {},
    }


def evidence_report(evidences: list[Any], questions: list[Any]) -> str:
    payload = {
        "evidences": [
            {
                "source_type": item.source_type,
                "source_ref": item.source_ref,
                "field": item.field,
                "value": item.value,
                "confidence": item.confidence,
                "status": item.status,
            }
            for item in evidences
        ],
        "questions": [
            {
                "step": item.step,
                "title": item.title,
                "severity": item.severity,
                "status": item.status,
            }
            for item in questions
        ],
    }
    return json.dumps(payload, indent=2)
